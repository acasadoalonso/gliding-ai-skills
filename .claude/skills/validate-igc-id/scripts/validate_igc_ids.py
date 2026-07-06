#!/usr/bin/env python3
"""Validate the "Igc id" column of an entry spreadsheet against the IGC
Ranking-list REST API (https://rankingdata.fai.org/rest/api/rlpilot?id=N).

Classifies every named person as VALID / WRONG / NOT SUPPLIED, writes a
markdown report, and saves a colour-coded copy of the workbook where the
"Igc id" cell is green (valid) or red (wrong).
"""

import argparse
import datetime as dt
import json
import re
import sys
import unicodedata
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

API_BASE = "https://rankingdata.fai.org/rest/api/rlpilot"
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

VALID, WRONG, NOT_SUPPLIED = "VALID", "WRONG", "NOT_SUPPLIED"


# Letters NFD decomposition doesn't strip (Ł has no combining mark)
_SPECIAL = str.maketrans({"ł": "l", "Ł": "l", "ø": "o", "Ø": "o", "đ": "d",
                          "Đ": "d", "ß": "ss", "æ": "ae", "Æ": "ae",
                          "œ": "oe", "Œ": "oe", "-": " ", "'": " "})


def norm(s):
    """Lowercase, accent-stripped, hyphen-free, squeezed string for names."""
    s = str(s or "").translate(_SPECIAL)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def clean_id(value):
    """Return the ID as a plain digit string, or None if not a clean number.

    Google-Sheets xlsx exports numbers as floats (5161.0) — accept those.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if re.fullmatch(r"\d+(\.0+)?", s):
        return s.split(".")[0]
    return None  # non-standard text (e.g. 'ARG-123') — not queryable


_cache = {}


def lookup(pilot_id):
    """Query the ranking list for one ID. Returns the pilot record or None."""
    if pilot_id in _cache:
        return _cache[pilot_id]
    url = f"{API_BASE}?id={urllib.parse.quote(pilot_id)}"
    try:
        with urllib.request.urlopen(url, timeout=30) as r:
            payload = json.loads(r.read().decode("utf-8"))
    except Exception as e:
        print(f"  ! API error for id {pilot_id}: {e}", file=sys.stderr)
        _cache[pilot_id] = {"error": str(e)}
        return _cache[pilot_id]
    data = payload.get("data")
    _cache[pilot_id] = data[0] if data else None
    return _cache[pilot_id]


def names_match(rec, first, last):
    """The registered pilot must be the same person as the spreadsheet row.

    Surnames must overlap (either contains the other, to allow compound
    surnames); if the surname alone is ambiguous the first name settles it.
    """
    rs, rf = norm(rec.get("surname")), norm(rec.get("firstname"))
    ls, lf = norm(last), norm(first)
    if not rs or not ls:
        return False
    if rs in ls or ls in rs:
        return True
    # Some sheets swap first/last name order
    return bool(rf and lf and (rs in lf or lf in rs) and (rf in ls or ls in rf))


def download_gsheet(url, dest):
    """Convert a Google-Sheets edit URL into an xlsx export and download it."""
    m = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", url)
    if not m:
        sys.exit(f"Cannot extract a spreadsheet id from URL: {url}")
    export = f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"
    req = urllib.request.Request(export, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as r, open(dest, "wb") as f:
        f.write(r.read())
    return dest


def find_columns(ws):
    """Locate the header row and the needed columns by header text."""
    wanted = {"igc id": "id", "first name": "first", "last name": "last",
              "country": "country"}
    for row in ws.iter_rows(min_row=1, max_row=10):
        cols = {}
        for cell in row:
            key = norm(cell.value)
            for header, tag in wanted.items():
                if key == header and tag not in cols:
                    cols[tag] = cell.column
        if "id" in cols and "first" in cols and "last" in cols:
            return row[0].row, cols
    sys.exit("Could not find a header row with 'Igc id', 'First name' and "
             "'Last name' columns in the first 10 rows.")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--excel", required=True,
                    help=".xlsx path, or a Google Sheets URL to download")
    ap.add_argument("--sheet", help="worksheet name (default: active sheet)")
    ap.add_argument("--report", help="report path (default: "
                    "reports/<stem>_igc_id_validation.md)")
    ap.add_argument("--out", help="coloured copy path (default: "
                    "<stem>_igc_id_validated.xlsx next to the input)")
    ap.add_argument("--generated", default=dt.date.today().isoformat(),
                    help="date stamped on the report")
    ap.add_argument("--title", help="report heading")
    args = ap.parse_args()

    src = args.excel
    if src.startswith("http"):
        dest = Path("Documents/igc_id_gsheet_download.xlsx")
        dest.parent.mkdir(exist_ok=True)
        src = str(download_gsheet(src, dest))
        print(f"Downloaded Google Sheet to {src}")
    src = Path(src)
    if not src.exists():
        sys.exit(f"File not found: {src}")

    wb = openpyxl.load_workbook(src)
    ws = wb[args.sheet] if args.sheet else wb.active
    header_row, cols = find_columns(ws)
    print(f"Sheet '{ws.title}': header row {header_row}, "
          f"Igc id column {openpyxl.utils.get_column_letter(cols['id'])}")

    results = []  # (row, first, last, country, raw_id, status, note)
    for row in ws.iter_rows(min_row=header_row + 1):
        r = row[0].row
        first = ws.cell(r, cols["first"]).value
        last = ws.cell(r, cols["last"]).value
        if not (first or last):
            continue  # country separator or empty slot
        country = ws.cell(r, cols["country"]).value if "country" in cols else ""
        raw = ws.cell(r, cols["id"]).value
        raw_str = "" if raw is None else str(raw).strip()

        if raw_str == "":
            results.append((r, first, last, country, "", NOT_SUPPLIED, ""))
            continue

        pid = clean_id(raw)
        if pid is None:
            results.append((r, first, last, country, raw_str, WRONG,
                            "not a plain ranking-list number"))
            continue

        rec = lookup(pid)
        if rec is None:
            results.append((r, first, last, country, pid, WRONG,
                            "id not found in the IGC ranking list"))
        elif "error" in rec:
            results.append((r, first, last, country, pid, WRONG,
                            f"unverifiable — API error ({rec['error']})"))
        elif norm(rec.get("surname")) == "blank":
            results.append((r, first, last, country, pid, WRONG,
                            "id exists but is an empty (Blank) profile — "
                            "cannot be confirmed as this pilot"))
        elif names_match(rec, first, last):
            note = ""
            nat = rec.get("nationality", "")
            if country and nat and norm(country) != norm(nat):
                note = f"nationality on record: {nat}"
            results.append((r, first, last, country, pid, VALID, note))
        else:
            registered = f"{rec.get('firstname','')} {rec.get('surname','')}".strip()
            results.append((r, first, last, country, pid, WRONG,
                            f"id belongs to {registered} "
                            f"({rec.get('nationality','?')})"))

    # ---- colour the Igc id cells on a copy ----
    out = Path(args.out) if args.out else src.with_name(
        src.stem + "_igc_id_validated.xlsx")
    # The source sheet may carry its own fills — overwrite every checked cell
    # so the colour always reflects this run's verdict (none = not supplied).
    for r, _f, _l, _c, _id, status, _n in results:
        cell = ws.cell(r, cols["id"])
        if status == VALID:
            cell.fill = GREEN
        elif status == WRONG:
            cell.fill = RED
        else:
            cell.fill = PatternFill(fill_type=None)
    wb.save(out)

    # ---- report ----
    report = Path(args.report) if args.report else \
        Path("reports") / f"{src.stem}_igc_id_validation.md"
    report.parent.mkdir(exist_ok=True)
    title = args.title or f"IGC Ranking ID validation — {src.stem}"
    buckets = {VALID: [], WRONG: [], NOT_SUPPLIED: []}
    for item in results:
        buckets[item[5]].append(item)

    def table(rows, with_note=True):
        head = "| Row | Name | Country | Igc id |" + (" Note |" if with_note else "")
        sep = "|---|---|---|---|" + ("---|" if with_note else "")
        lines = [head, sep]
        for r, f, l, c, i, _s, n in rows:
            base = f"| {r} | {f} {l} | {c or ''} | {i or '—'} |"
            lines.append(base + (f" {n} |" if with_note else ""))
        return "\n".join(lines)

    md = [f"# {title}", "",
          f"- Generated: {args.generated}",
          f"- Source: {src.name} (sheet '{ws.title}')",
          f"- Checked against: {API_BASE}", "",
          "## 1. Summary", "",
          "| Outcome | Count |", "|---|---|",
          f"| ✅ Valid | {len(buckets[VALID])} |",
          f"| ❌ Wrong | {len(buckets[WRONG])} |",
          f"| ⚠️ Not supplied | {len(buckets[NOT_SUPPLIED])} |",
          f"| Total checked | {len(results)} |", "",
          "## 2.1 Valid IDs", "", table(buckets[VALID]), "",
          "## 2.2 Wrong IDs", "", table(buckets[WRONG]), "",
          "## 2.3 Not supplied", "", table(buckets[NOT_SUPPLIED], with_note=False), ""]
    report.write_text("\n".join(md), encoding="utf-8")

    print(f"\nValid: {len(buckets[VALID])}  Wrong: {len(buckets[WRONG])}  "
          f"Not supplied: {len(buckets[NOT_SUPPLIED])}")
    print(f"Report:  {report}")
    print(f"Workbook copy: {out}")


if __name__ == "__main__":
    main()
