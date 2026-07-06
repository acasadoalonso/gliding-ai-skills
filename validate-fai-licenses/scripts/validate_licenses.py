#!/usr/bin/python3
# -*- coding: UTF-8 -*-
"""
Validate the FAI sporting licenses listed in a competition-entry .xlsx against
the official FAI extranet records, then colour the entry rows by outcome.

Logic follows the memory file validate-fai-sporting-license-for-pilots.md:
  * fetch every Gliding and Universal licence for the pilot's country (FAI
    extranet API, paginated 100/page) — a Universal licence covers all FAI
    sports, so it is as valid as a Gliding one,
  * match the provided licence number numerically against `idlicencee`,
  * when the number is wrong / non-standard / missing, fall back to a
    name match (surname + given name) to recover the correct number.

Outcomes and row colours:
  VALID       provided number matches an FAI record      -> green
  NAME_MATCH  number wrong or missing, found by name      -> blue
  INVALID     number provided but no number/name match    -> red
  MISSING     no number provided and no name match        -> (uncoloured,
              except CPT/captain rows, which are painted red: a captain
              needs a licence, so nothing found is a failed validation)

The entry sheet is read by column *header* (first_name, last_name,
country_code, fai_licence_number), so the same script works for any
SoaringSpot-style entry export.
"""

import argparse
import datetime
import json
import os
import re
import sys
import unicodedata
import urllib.request

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ISO 3166 alpha-3 -> IOC code, only where they differ. The FAI API expects the
# IOC code (e.g. GER, not DEU). Codes not listed are assumed identical.
ISO3_TO_IOC = {
    "DEU": "GER", "NLD": "NED", "CHE": "SUI", "DNK": "DEN", "HRV": "CRO",
    "PRT": "POR", "GRC": "GRE", "SVN": "SLO", "LVA": "LAT", "BGR": "BUL",
    "CHL": "CHI", "MYS": "MAS", "IDN": "INA", "IRN": "IRI", "ZAF": "RSA",
    "MCO": "MON", "PHL": "PHI", "ARE": "UAE", "SAU": "KSA", "OMN": "OMA",
    "KWT": "KUW", "LBN": "LIB", "VNM": "VIE", "PRY": "PAR", "URY": "URU",
    "NGA": "NGR", "ZWE": "ZIM", "ZMB": "ZAM", "TZA": "TAN", "AGO": "ANG",
    "DZA": "ALG", "BGD": "BAN", "NPL": "NEP", "LKA": "SRI", "GTM": "GUA",
    "HND": "HON", "CRI": "CRC", "SLV": "ESA", "PRI": "PUR",
}

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")   # VALID
FILL_BLUE = PatternFill("solid", fgColor="BDD7EE")    # NAME_MATCH
FILL_RED = PatternFill("solid", fgColor="FFC7CE")     # INVALID


# --------------------------------------------------------------------------- #
# FAI password
# --------------------------------------------------------------------------- #
def resolve_password(cli_value, config_dir):
    """Password precedence: --fai-password, $FAIPWD env, then tools/config.py."""
    if cli_value:
        return cli_value
    if os.environ.get("FAIPWD"):
        return os.environ["FAIPWD"]
    if config_dir and os.path.isdir(config_dir):
        sys.path.insert(0, config_dir)
        try:
            import config  # noqa: E402
            return config.FAIPWD
        except Exception:
            pass
    sys.exit("FAI password not found. Pass --fai-password, set $FAIPWD, "
             "or ensure config.py (with FAIPWD) is in --config-dir.")


# --------------------------------------------------------------------------- #
# FAI extranet API
# --------------------------------------------------------------------------- #
def get_licenses_per_country(country, password, cache):
    """All Gliding and Universal licences for one IOC country, paginated.
    Cached per country. Universal licences cover every FAI sport, but the
    discipline=Gliding query never returns them, so each discipline is
    fetched separately."""
    if country in cache:
        return cache[country]
    out = []
    for sport in ("Gliding", "Universal"):
        start = 0
        nl = 100
        while nl == 100:
            url = ("https://extranet.fai.org/api/v1/licences"
                   "?auth_username=FAIOrganizer&auth_password=" + password +
                   "&discipline=" + sport + "&country=" + country +
                   "&limit_length=100&limit_start=" + str(start))
            with urllib.request.urlopen(url, timeout=60) as j:
                batch = json.loads(j.read().decode("UTF-8"))
            nl = len(batch)
            out.extend(l for l in batch if l.get("Sport") == sport)
            start += nl
    cache[country] = out
    return out


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def norm(s):
    """Lower-case, accent-stripped, whitespace-collapsed name for comparison."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def clean_number(raw):
    """Return the licence as a plain integer string if it *is* one clean number.

    A clean number (e.g. '162862', '162862.0') can be matched directly and, if
    found, counts as VALID. Compound / country-prefixed formats like
    'GER-4264 ID-129831' or 'POL-110/06' return None so they route to the name
    match and are reported as corrections.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if re.fullmatch(r"\d+(\.0+)?", s):
        return str(int(float(s)))
    return None


def name_match(first, last, lics):
    """Find a licence record by surname (+ given name to disambiguate)."""
    last_n = norm(last)
    first_n = norm(first)
    cands = [l for l in lics if norm(l.get("surname_lip")) == last_n]
    if not cands:
        return None
    if len(cands) == 1:
        return cands[0]
    # Multiple people share the surname: require a given-name signal.
    for l in cands:
        g = norm(l.get("givenname_lip"))
        if not g or not first_n:
            continue
        if g == first_n or g.split(" ")[0] == first_n.split(" ")[0] \
                or first_n.startswith(g) or g.startswith(first_n):
            return l
    return None


def find_headers(ws):
    """Locate the header row and the columns we need, by header text."""
    aliases = {
        "first_name": {"first_name", "firstname", "given_name", "givenname"},
        "last_name": {"last_name", "lastname", "surname", "family_name", "name"},
        "country_code": {"country_code", "country", "nationality", "nation"},
        "fai": {"fai_licence_number", "fai_license_number",
                "fai_licence", "fai_license", "licence_number",
                "fai_sporting_licence_number", "fai_sporting_license_number",
                "sporting_licence_number", "sporting_license_number"},
    }
    for r in range(1, min(ws.max_row, 10) + 1):
        found = {}
        for c in range(1, ws.max_column + 1):
            key = norm(ws.cell(r, c).value).replace(" ", "_")
            for field, names in aliases.items():
                if key in names and field not in found:
                    found[field] = c
        if "first_name" in found and "last_name" in found:
            return r, found
    sys.exit("Could not find a header row with first_name/last_name columns.")


def find_role_col(ws, header_row):
    """Column holding the EGC role marker ('CPT' for captains, running numbers
    for pilots). It has no header, so we detect it by scanning data rows for a
    'CPT' value. Returns the column index, or None if the form has no roles."""
    for c in range(1, min(ws.max_column, 4) + 1):
        for r in range(header_row + 1, min(ws.max_row, header_row + 200) + 1):
            if norm(ws.cell(r, c).value) == "cpt":
                return c
    return None


# --------------------------------------------------------------------------- #
# Core
# --------------------------------------------------------------------------- #
def validate(ws, header_row, cols, password, cache, role_col=None):
    """Classify every data row. Returns list of per-pilot result dicts."""
    results = []
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(r, cols["first_name"]).value
        last = ws.cell(r, cols["last_name"]).value
        if not first and not last:
            continue
        iso = str(ws.cell(r, cols["country_code"]).value or "").strip().upper() \
            if "country_code" in cols else ""
        ioc = ISO3_TO_IOC.get(iso, iso)
        raw = ws.cell(r, cols["fai"]).value if "fai" in cols else None
        raw_str = "" if raw is None else str(raw).strip()
        is_captain = role_col is not None and norm(ws.cell(r, role_col).value) == "cpt"

        # Skip section-title rows (e.g. "Club Class"): a value only in the
        # surname column, with no first name, country, or licence number.
        if not first and not iso and not raw_str:
            continue

        res = {
            "row": r, "first": str(first or "").strip(),
            "last": str(last or "").strip(), "iso": iso, "ioc": ioc,
            "provided": raw_str, "status": None, "correct": None,
            "valid_until": None, "expired": None,
            "is_captain": is_captain, "resolved": None,
        }

        if not ioc:
            res["status"] = "INVALID" if raw_str else "MISSING"
            res["note"] = "no country code"
            results.append(res)
            continue

        try:
            lics = get_licenses_per_country(ioc, password, cache)
        except Exception as e:  # country fetch failed -> cannot verify
            res["status"] = "INVALID"
            res["note"] = f"FAI lookup failed for {ioc}: {e}"
            results.append(res)
            continue

        # 1) direct number match -> VALID
        cn = clean_number(raw)
        rec = None
        if cn is not None:
            rec = next((l for l in lics if str(l.get("idlicencee")) == cn), None)
            if rec:
                res["status"] = "VALID"

        # 2) fall back to name match -> NAME_MATCH (correction / recovery)
        if res["status"] is None:
            rec = name_match(res["first"], res["last"], lics)
            if rec:
                res["status"] = "NAME_MATCH"
                res["correct"] = rec.get("idlicencee")
            else:
                res["status"] = "INVALID" if raw_str else "MISSING"

        if rec:
            res["valid_until"] = rec.get("validuntil_lic")
            res["expired"] = bool(rec.get("is_expired"))
            res["resolved"] = str(rec.get("idlicencee"))
        results.append(res)
    return results


def colour_rows(ws, results, ncols):
    fill_for = {"VALID": FILL_GREEN, "NAME_MATCH": FILL_BLUE, "INVALID": FILL_RED}
    for res in results:
        fill = fill_for.get(res["status"])
        if not fill and res["status"] == "MISSING" and res.get("is_captain"):
            # A captain needs a licence: nothing found means the row failed
            # validation, so paint it red rather than leaving it uncoloured.
            fill = FILL_RED
        if not fill:
            continue
        for c in range(1, ncols + 1):
            ws.cell(res["row"], c).fill = fill


def fill_captain_licences(ws, results, fai_col):
    """Write the resolved FAI number into a captain's licence cell.

    Team captains often leave the licence column blank or enter a non-standard
    value, but a captain still needs a valid licence — so whenever validation
    resolved a real FAI record for a CPT row, write that number back into the
    entry sheet. Returns the list of rows updated (for reporting)."""
    updated = []
    for res in results:
        if not res.get("is_captain") or not res.get("resolved"):
            continue
        if res["provided"] == res["resolved"]:
            continue  # already correct, leave the cell untouched
        ws.cell(res["row"], fai_col).value = res["resolved"]
        updated.append(res)
    return updated


def write_report(path, results, source_name, title, when, captain_fill=True):
    valid = [r for r in results if r["status"] == "VALID"]
    named = [r for r in results if r["status"] == "NAME_MATCH"]
    invalid = [r for r in results if r["status"] == "INVALID"]
    missing = [r for r in results if r["status"] == "MISSING"]
    captains_filled = [
        r for r in results
        if r.get("is_captain") and r.get("resolved")
        and r["provided"] != r["resolved"]
    ] if captain_fill else []

    def name_of(r):
        return f"{r['first']} {r['last']}".strip()

    lines = []
    lines.append(f"# {title}")
    lines.append("")
    lines.append(f"**Date of Validation:** {when}")
    lines.append(f"**Source Data:** {source_name}")
    lines.append(f"**Total Pilots Processed:** {len(results)}")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 1. Executive Summary")
    lines.append("")
    lines.append("| Status | Count | Description |")
    lines.append("| :--- | :--- | :--- |")
    lines.append(f"| **VALID** | {len(valid)} | Licence number matched an active FAI record. |")
    lines.append(f"| **FOUND BY NAME** | {len(named)} | Number wrong/missing, pilot found via name match. |")
    lines.append(f"| **INVALID** | {len(invalid)} | Provided number not found in the national records. |")
    lines.append(f"| **NO LICENSE PROVIDED** | {len(missing)} | No number listed and no name match found. |")
    lines.append(f"| **TOTAL** | **{len(results)}** | |")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 2. Detailed Findings")
    lines.append("")

    lines.append("### 2.1 Validated Licenses")
    lines.append("The following pilots have verified, active FAI sporting licenses:")
    lines.append("")
    for r in sorted(valid, key=lambda x: (x["ioc"], x["last"])):
        exp = " — **EXPIRED**" if r["expired"] else ""
        vu = f", valid until {r['valid_until']}" if r["valid_until"] else ""
        lines.append(f"* {name_of(r)} ({r['ioc']}) - License: {r['provided']}{vu}{exp}")
    lines.append("")

    lines.append("### 2.2 Corrected via Name Match")
    lines.append("Number was non-standard, incorrect, or absent; pilot identified by name:")
    lines.append("")
    for r in sorted(named, key=lambda x: (x["ioc"], x["last"])):
        exp = " — **EXPIRED**" if r["expired"] else ""
        if r["provided"]:
            lines.append(f"* {name_of(r)} ({r['ioc']}): Provided `{r['provided']}` -> Correct: `{r['correct']}`{exp}")
        else:
            lines.append(f"* {name_of(r)} ({r['ioc']}): License: {r['correct']} (no number on entry){exp}")
    lines.append("")

    lines.append("### 2.3 Invalid/Unverifiable Licenses")
    lines.append("Provided license numbers that could not be found in national FAI records:")
    lines.append("")
    for r in sorted(invalid, key=lambda x: (x["ioc"], x["last"])):
        note = f" ({r['note']})" if r.get("note") else ""
        lines.append(f"* {name_of(r)} ({r['ioc']}) — License {r['provided'] or 'N/A'}{note}")
    lines.append("")

    lines.append("### 2.4 Missing Information")
    lines.append("No license number provided and no name match:")
    lines.append("")
    for r in sorted(missing, key=lambda x: (x["ioc"], x["last"])):
        lines.append(f"* {name_of(r)} ({r['ioc']})")
    lines.append("")

    if captains_filled:
        lines.append("### 2.5 Captain Licence Numbers Written Back")
        lines.append("Team-captain (CPT) rows whose licence cell was blank or "
                     "non-standard; the resolved FAI number was written into the "
                     "entry sheet:")
        lines.append("")
        for r in sorted(captains_filled, key=lambda x: (x["ioc"], x["last"])):
            shown = f"`{r['provided']}`" if r["provided"] else "(blank)"
            lines.append(f"* {name_of(r)} ({r['ioc']}): {shown} -> `{r['resolved']}`")
        lines.append("")

    with open(path, "w") as f:
        f.write("\n".join(lines))


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Validate FAI licenses in an .xlsx entry list and colour the rows.")
    ap.add_argument("--excel", required=True, help="Path to the entry .xlsx")
    ap.add_argument("--out", help="Coloured workbook output (default: overwrite --excel)")
    ap.add_argument("--report", help="Report path (default: reports/<stem>_fai_license_validation.txt)")
    ap.add_argument("--sheet", help="Worksheet name (default: active sheet)")
    ap.add_argument("--title", help="Report title")
    ap.add_argument("--no-color", action="store_true", help="Validate + report only, do not colour the workbook")
    ap.add_argument("--no-captain-fill", action="store_true", help="Do not write resolved licence numbers into blank/wrong CPT (captain) cells")
    ap.add_argument("--fai-password", help="FAI extranet password (base64), overrides env/config")
    ap.add_argument("--config-dir", default="/home/angel/tools", help="Dir containing config.py with FAIPWD")
    ap.add_argument("--generated", help="Validation date for the report (YYYY-MM-DD)")
    args = ap.parse_args()

    password = resolve_password(args.fai_password, args.config_dir)
    wb = load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active

    header_row, cols = find_headers(ws)
    role_col = find_role_col(ws, header_row)
    cache = {}
    results = validate(ws, header_row, cols, password, cache, role_col)

    stem = os.path.splitext(os.path.basename(args.excel))[0]
    report = args.report or os.path.join("reports", f"{stem}_fai_license_validation.txt")
    os.makedirs(os.path.dirname(os.path.abspath(report)), exist_ok=True)
    title = args.title or f"{stem} FAI License Validation Report"
    when = args.generated or datetime.date.today().isoformat()
    captain_fill = not args.no_captain_fill and not args.no_color
    write_report(report, results, os.path.basename(args.excel), title, when,
                 captain_fill=captain_fill)

    captain_updates = []
    if not args.no_color:
        colour_rows(ws, results, ws.max_column)
        if "fai" in cols and not args.no_captain_fill:
            captain_updates = fill_captain_licences(ws, results, cols["fai"])
        out = args.out or args.excel
        wb.save(out)

    counts = {}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    print("Pilots:", len(results))
    for k in ("VALID", "NAME_MATCH", "INVALID", "MISSING"):
        print(f"  {k}: {counts.get(k, 0)}")
    print("Report:", report)
    if not args.no_color:
        print("Workbook:", args.out or args.excel, "(green=valid, blue=name-match, red=invalid)")
        print(f"Captain licence cells filled: {len(captain_updates)}")
        for r in captain_updates:
            shown = r["provided"] or "(blank)"
            print(f"  row {r['row']} {r['first']} {r['last']} ({r['ioc']}): {shown} -> {r['resolved']}")


if __name__ == "__main__":
    main()
